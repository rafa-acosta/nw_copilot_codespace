"""
Microsoft Excel file loader.

Handles loading and preprocessing of XLSX files.
"""

from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from .base import BaseDataLoader, IngestedData, IngestSource
from .cleaning import TextCleaner
from .utils import validate_file_exists, validate_file_extension


class ExcelLoader(BaseDataLoader):
    """
    Loader for Microsoft Excel (.xlsx) files.
    
    Extracts text from all sheets and cells. Applies cleaning
    to prepare content for RAG processing.
    
    Example:
        >>> loader = ExcelLoader()
        >>> data = loader.load('spreadsheet.xlsx')
        >>> print(data.content[:100])
    """
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    
    def __init__(self, cleaner: Optional[TextCleaner] = None):
        """
        Initialize ExcelLoader.
        
        Args:
            cleaner: Optional TextCleaner instance.
            
        Raises:
            ImportError: If openpyxl is not installed.
        """
        if load_workbook is None:
            raise ImportError(
                "openpyxl is required for Excel loading. "
                "Install with: pip install openpyxl"
            )
        
        self.cleaner = cleaner or TextCleaner.for_structured_data()
    
    def validate_source(self, source: str) -> bool:
        """
        Validate that source is a valid Excel file.
        
        Args:
            source: File path to validate.
            
        Returns:
            True if valid Excel file, False otherwise.
        """
        if not validate_file_exists(source):
            return False
        return validate_file_extension(source, self.SUPPORTED_EXTENSIONS)
    
    def load(self, source: str, **kwargs) -> IngestedData:
        """
        Load and extract data from an Excel file.
        
        Args:
            source: Path to the Excel file.
            **kwargs: Optional parameters:
                - include_empty_cells (bool): Include empty cells (default: False).
                - sheet_names (list): Specific sheets to load (default: all).
                
        Returns:
            IngestedData with extracted and cleaned text.
            
        Raises:
            FileNotFoundError: If file does not exist.
            ValueError: If file is not a valid Excel file.
        """
        if not self.validate_source(source):
            raise ValueError(f"Invalid Excel file: {source}")
        
        include_empty = kwargs.get('include_empty_cells', False)
        sheet_names = kwargs.get('sheet_names', None)
        metadata = {}
        
        try:
            workbook = load_workbook(source)
            text_parts = []
            
            # Get sheets to process
            sheets = sheet_names if sheet_names else workbook.sheetnames
            
            for sheet_name in sheets:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                
                # Extract cell values
                for row in sheet.iter_rows(values_only=True):
                    row_values = []
                    for cell_value in row:
                        if cell_value is not None:
                            row_values.append(str(cell_value).strip())
                        elif include_empty:
                            row_values.append("[EMPTY]")
                    
                    if row_values:
                        text_parts.append(' | '.join(row_values))
            
            raw_text = '\n'.join(text_parts)
            metadata['sheet_count'] = len(workbook.sheetnames)
            metadata['sheets'] = workbook.sheetnames
            
        except Exception as e:
            raise ValueError(f"Failed to read Excel file {source}: {str(e)}")
        
        # Clean the extracted text
        cleaned_text = self.cleaner.clean(raw_text)
        
        # Create metadata
        ingest_source = self._create_ingest_source(
            source=source,
            source_type='excel',
            metadata=metadata
        )
        
        return IngestedData(content=cleaned_text, source=ingest_source)
